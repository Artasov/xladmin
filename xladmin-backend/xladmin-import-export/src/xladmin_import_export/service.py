from __future__ import annotations

import csv
import json
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO, StringIO
from typing import Any
from uuid import UUID, uuid4

from fastapi import HTTPException, UploadFile, status
from openpyxl import Workbook, load_workbook
from sqlalchemy import inspect as sa_inspect
from sqlalchemy.ext.asyncio import AsyncSession
from xladmin.config import AdminModelConfig
from xladmin.introspection_fields import (
    get_all_field_names,
    get_column_names,
    get_mapper_field_names,
    get_pk_field_name,
    get_relationship_names,
    pk_is_generated,
)
from xladmin.introspection_values import convert_value_for_column, get_field_value
from xladmin.router_mutations import apply_payload_to_item
from xladmin.router_queries import (
    apply_list_filters,
    apply_ordering,
    apply_search,
    build_model_query,
    get_items_by_ids,
)
from xladmin.serializer import serialize_scalar

from xladmin_import_export.config import ImportConflictMode, ImportExportConfig, ImportExportFormat
from xladmin_import_export.schemas import (
    ExportRequestPayload,
    ImportCommitResponse,
    ImportPreviewItem,
    ImportValidationErrorPayload,
    ImportValidationResponse,
    ImportValidationSummary,
)


@dataclass(slots=True)
class ExportArtifact:
    filename: str
    content_type: str
    content: bytes


def get_import_export_config(model_config: AdminModelConfig) -> ImportExportConfig | None:
    config = getattr(model_config, "import_export", None)
    if config is None:
        return None
    if not isinstance(config, ImportExportConfig):
        raise TypeError("model_config.import_export must be ImportExportConfig.")
    return config


def resolve_export_fields(model_config: AdminModelConfig, config: ImportExportConfig) -> list[str]:
    if config.export_fields is not None:
        return list(dict.fromkeys(config.export_fields))

    fields: list[str] = []
    for field_name in get_all_field_names(model_config):
        field_config = model_config.get_field_config(field_name)
        if field_config.input_kind == "password":
            continue
        fields.append(field_name)
    return list(dict.fromkeys(fields))


def resolve_default_export_fields(model_config: AdminModelConfig, config: ImportExportConfig) -> list[str]:
    if config.default_export_fields is not None:
        return [field_name for field_name in dict.fromkeys(config.default_export_fields) if field_name]
    return resolve_export_fields(model_config, config)


def resolve_import_fields(model_config: AdminModelConfig, config: ImportExportConfig) -> list[str]:
    if config.import_fields is not None:
        return list(dict.fromkeys(config.import_fields))

    column_names = set(get_column_names(model_config))
    relationship_names = set(get_relationship_names(model_config))
    return [
        field_name for field_name in get_mapper_field_names(model_config)
        if field_name in column_names or field_name in relationship_names
    ]


def resolve_default_import_fields(model_config: AdminModelConfig, config: ImportExportConfig) -> list[str]:
    if config.default_import_fields is not None:
        return [field_name for field_name in dict.fromkeys(config.default_import_fields) if field_name]
    return resolve_import_fields(model_config, config)


async def resolve_export_items(
    session: AsyncSession,
    model_config: AdminModelConfig,
    user: Any,
    payload: ExportRequestPayload,
) -> list[Any]:
    if payload.select_all:
        scope = payload.selection_scope
        query = await build_model_query(session, model_config, user, mode="detail")
        query = apply_search(model_config, query, scope.q if scope is not None else None, session)
        query = await apply_list_filters(
            model_config,
            query,
            scope.filters if scope is not None else {},
            session,
            user,
        )
        query = apply_ordering(model_config, query, scope.sort if scope is not None else None)
        return list((await session.execute(query)).scalars().unique())

    if payload.ids:
        return await get_items_by_ids(session, model_config, payload.ids, user, mode="detail")

    query = await build_model_query(session, model_config, user, mode="detail")
    search_value = payload.selection_scope.q if payload.selection_scope is not None else None
    sort_value = payload.selection_scope.sort if payload.selection_scope is not None else None
    filter_values = payload.selection_scope.filters if payload.selection_scope is not None else {}
    query = apply_search(model_config, query, search_value, session)
    query = await apply_list_filters(
        model_config,
        query,
        filter_values,
        session,
        user,
    )
    query = apply_ordering(model_config, query, sort_value)
    return list((await session.execute(query)).scalars().unique())


def build_export_artifact(
    model_config: AdminModelConfig,
    items: list[Any],
    fields: list[str],
    export_format: ImportExportFormat,
) -> ExportArtifact:
    rows = [serialize_export_row(model_config, item, fields, export_format) for item in items]
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    filename_base = f"{model_config.slug}-export-{timestamp}"

    if export_format == "json":
        return ExportArtifact(
            filename=f"{filename_base}.json",
            content_type="application/json",
            content=json.dumps(rows, ensure_ascii=False, indent=2).encode("utf-8"),
        )
    if export_format == "csv":
        buffer = StringIO()
        writer = csv.DictWriter(buffer, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)
        return ExportArtifact(
            filename=f"{filename_base}.csv",
            content_type="text/csv; charset=utf-8",
            content=buffer.getvalue().encode("utf-8-sig"),
        )
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = model_config.slug
    worksheet.append(fields)
    for row in rows:
        worksheet.append([row.get(field_name) for field_name in fields])
    output = BytesIO()
    workbook.save(output)
    return ExportArtifact(
        filename=f"{filename_base}.xlsx",
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        content=output.getvalue(),
    )


def serialize_export_row(
    model_config: AdminModelConfig,
    item: Any,
    fields: list[str],
    export_format: ImportExportFormat,
) -> dict[str, Any]:
    row: dict[str, Any] = {}
    for field_name in fields:
        row[field_name] = normalize_export_value(get_field_value(model_config, item, field_name), export_format)
    return row


def normalize_export_value(value: Any, export_format: ImportExportFormat) -> Any:
    scalar_value = serialize_scalar(value)
    if export_format == "json":
        return scalar_value
    if isinstance(scalar_value, list | dict):
        return json.dumps(scalar_value, ensure_ascii=False)
    if isinstance(scalar_value, bool):
        return "true" if scalar_value else "false"
    return scalar_value


async def read_import_rows(upload_file: UploadFile, import_format: ImportExportFormat) -> list[dict[str, Any]]:
    file_bytes = await upload_file.read()
    if import_format == "json":
        payload = json.loads(file_bytes.decode("utf-8-sig"))
        if not isinstance(payload, list) or any(not isinstance(item, dict) for item in payload):
            raise ValueError("JSON file must contain an array of objects.")
        return [dict(item) for item in payload]
    if import_format == "csv":
        text = file_bytes.decode("utf-8-sig")
        return [dict(item) for item in csv.DictReader(StringIO(text))]
    workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    return [
        {
            headers[index]: row[index]
            for index in range(len(headers))
            if headers[index]
        }
        for row in rows[1:]
        if any(value not in (None, "") for value in row)
    ]


async def validate_import_rows(
    session: AsyncSession,
    model_config: AdminModelConfig,
    rows: list[dict[str, Any]],
    fields: list[str],
    conflict_mode: ImportConflictMode,
) -> ImportValidationResponse:
    preview = await process_import_rows(
        session,
        model_config,
        rows,
        fields,
        conflict_mode,
        commit=False,
    )
    await session.rollback()
    return preview


async def commit_import_rows(
    session: AsyncSession,
    model_config: AdminModelConfig,
    rows: list[dict[str, Any]],
    fields: list[str],
    conflict_mode: ImportConflictMode,
) -> ImportCommitResponse:
    preview = await process_import_rows(
        session,
        model_config,
        rows,
        fields,
        conflict_mode,
        commit=True,
    )
    return ImportCommitResponse(
        created=preview.summary.create,
        updated=preview.summary.update,
        skipped=preview.summary.skip,
    )


async def process_import_rows(
    session: AsyncSession,
    model_config: AdminModelConfig,
    rows: list[dict[str, Any]],
    fields: list[str],
    conflict_mode: ImportConflictMode,
    *,
    commit: bool,
) -> ImportValidationResponse:
    mapper = sa_inspect(model_config.model)
    pk_field = get_pk_field_name(model_config)
    pk_column = mapper.columns[pk_field]
    import_export_config = get_import_export_config(model_config) or ImportExportConfig()
    available_import_fields = set(resolve_import_fields(model_config, import_export_config))
    normalized_fields = [field_name for field_name in dict.fromkeys(fields) if field_name in available_import_fields]
    if not normalized_fields:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No valid import fields were selected.")
    errors: list[ImportValidationErrorPayload] = []
    created_preview: list[ImportPreviewItem] = []
    updated_preview: list[ImportPreviewItem] = []
    skipped_preview: list[ImportPreviewItem] = []
    created = 0
    updated = 0
    skipped = 0

    for row_index, raw_row in enumerate(rows, start=2):
        try:
            normalized_row = normalize_import_row(model_config, raw_row, normalized_fields)
            raw_pk_value = normalized_row.get(pk_field)
            existing_item = None
            normalized_pk_value = None
            if raw_pk_value not in (None, ""):
                normalized_pk_value = convert_value_for_column(pk_column, raw_pk_value)
                existing_item = await session.get(model_config.model, normalized_pk_value)

            if existing_item is not None:
                if conflict_mode == "skip_existing":
                    skipped += 1
                    skipped_preview.append(
                        ImportPreviewItem(
                            row_number=row_index,
                            label=build_preview_label(model_config, normalized_row, row_index),
                        )
                    )
                    continue
                if conflict_mode == "update_existing":
                    payload = {key: value for key, value in normalized_row.items() if key != pk_field}
                    await apply_payload_to_item(
                        session,
                        model_config,
                        existing_item,
                        payload,
                        mode="update",
                        allowed_fields=set(normalized_fields),
                    )
                    updated += 1
                    updated_preview.append(
                        ImportPreviewItem(
                            row_number=row_index,
                            label=build_preview_label(model_config, normalized_row, row_index),
                        )
                    )
                    continue

            item = model_config.model()
            payload = dict(normalized_row)
            if conflict_mode == "auto_generate_pk":
                payload.pop(pk_field, None)
                assign_generated_pk(item, pk_field, pk_column)
            elif normalized_pk_value is not None:
                setattr(item, pk_field, normalized_pk_value)
                payload.pop(pk_field, None)
            elif supports_auto_generate_pk(model_config):
                assign_generated_pk(item, pk_field, pk_column)
            await apply_payload_to_item(
                session,
                model_config,
                item,
                payload,
                mode="create",
                allowed_fields=set(normalized_fields),
            )
            missing_required_fields = get_missing_required_columns_for_item(model_config, item)
            if missing_required_fields:
                raise ValueError(
                    f"Missing required fields for create: {', '.join(missing_required_fields)}."
                )
            session.add(item)
            created += 1
            created_preview.append(
                ImportPreviewItem(
                    row_number=row_index,
                    label=build_preview_label(model_config, normalized_row, row_index),
                )
            )
        except HTTPException as exc:
            errors.append(ImportValidationErrorPayload(row_number=row_index, message=str(exc.detail)))
        except Exception as exc:  # noqa: BLE001
            errors.append(ImportValidationErrorPayload(row_number=row_index, message=str(exc)))

    if errors and commit:
        await session.rollback()
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail={
                "summary": {
                    "total_rows": len(rows),
                    "create": created,
                    "update": updated,
                    "skip": skipped,
                    "errors": len(errors),
                },
                "errors": [error.model_dump() for error in errors],
            },
        )

    if commit:
        await session.commit()

    return ImportValidationResponse(
        summary=ImportValidationSummary(
            total_rows=len(rows),
            create=created,
            update=updated,
            skip=skipped,
            errors=len(errors),
        ),
        created_preview=created_preview[:20],
        updated_preview=updated_preview[:20],
        skipped_preview=skipped_preview[:20],
        errors=errors[:50],
    )


def normalize_import_row(
    model_config: AdminModelConfig,
    raw_row: dict[str, Any],
    fields: list[str],
) -> dict[str, Any]:
    normalized_row: dict[str, Any] = {}
    for field_name in fields:
        if field_name not in raw_row:
            continue
        value = raw_row[field_name]
        if isinstance(value, str):
            value = value.strip()
        if value in ("", None):
            continue
        field_config = model_config.get_field_config(field_name)
        if field_config.input_kind == "relation-multiple" and isinstance(value, str):
            normalized_row[field_name] = parse_many_relation_value(value)
            continue
        if field_config.input_kind == "json" and isinstance(value, str):
            normalized_row[field_name] = json.loads(value)
            continue
        normalized_row[field_name] = value
    return normalized_row


def parse_many_relation_value(value: str) -> list[Any]:
    normalized_value = value.strip()
    if not normalized_value:
        return []
    if normalized_value.startswith("["):
        parsed = json.loads(normalized_value)
        if isinstance(parsed, list):
            return parsed
    return [item.strip() for item in normalized_value.split(",") if item.strip()]


def assign_generated_pk(item: Any, pk_field: str, pk_column: Any) -> None:
    try:
        python_type = pk_column.type.python_type
    except NotImplementedError:
        python_type = None
    if python_type is UUID:
        setattr(item, pk_field, uuid4())


def get_missing_required_columns_for_item(
    model_config: AdminModelConfig,
    item: Any,
) -> list[str]:
    mapper = sa_inspect(model_config.model)
    pk_field = get_pk_field_name(model_config)
    missing_fields: list[str] = []

    for column in mapper.columns:
        field_name = column.key
        if field_name == pk_field and supports_auto_generate_pk(model_config):
            continue
        if not is_required_create_column(column):
            continue
        if getattr(item, field_name, None) is not None:
            continue

        missing_fields.append(field_name)

    return missing_fields


def is_required_create_column(column: Any) -> bool:
    if bool(column.nullable):
        return False
    if column.default is not None or column.server_default is not None:
        return False
    if bool(column.primary_key) and pk_is_generated(column):
        return False
    return True


def build_preview_label(model_config: AdminModelConfig, row: dict[str, Any], row_number: int) -> str:
    display_field = model_config.display_field
    if display_field and row.get(display_field) not in (None, ""):
        return str(row[display_field])
    pk_field = get_pk_field_name(model_config)
    if row.get(pk_field) not in (None, ""):
        return str(row[pk_field])
    return f"Row {row_number}"


def resolve_pk_type_label(model_config: AdminModelConfig) -> str:
    pk_field = get_pk_field_name(model_config)
    pk_column = sa_inspect(model_config.model).columns[pk_field]
    try:
        python_type = pk_column.type.python_type
    except NotImplementedError:
        return "unknown"
    if python_type is int:
        return "int"
    if python_type is UUID:
        return "uuid"
    if python_type is str:
        return "string"
    return getattr(python_type, "__name__", "unknown")


def supports_auto_generate_pk(model_config: AdminModelConfig) -> bool:
    pk_field = get_pk_field_name(model_config)
    pk_column = sa_inspect(model_config.model).columns[pk_field]
    try:
        python_type = pk_column.type.python_type
    except NotImplementedError:
        python_type = None
    if python_type is UUID:
        return True
    return pk_is_generated(pk_column)
